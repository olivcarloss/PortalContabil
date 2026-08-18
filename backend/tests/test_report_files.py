from openpyxl import load_workbook

from app.core.report_files import build_csv, build_pdf, build_xlsx


def test_build_csv_includes_header_and_rows():
    out = build_csv(["Nome", "Valor"], [["Alice", "10"], ["Bob", "20"]])
    text = out.decode("utf-8-sig")
    assert "Nome;Valor" in text
    assert "Alice;10" in text
    assert "Bob;20" in text


def test_build_xlsx_includes_header_and_rows():
    out = build_xlsx(["Nome", "Valor"], [["Alice", "10"]])
    wb = load_workbook(filename=__import__("io").BytesIO(out))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Nome", "Valor"]
    assert [c.value for c in ws[2]] == ["Alice", "10"]


def test_build_pdf_returns_nonempty_bytes():
    out = build_pdf("Relatório Teste", ["Nome"], [["Alice"]], "https://example.com/logo.svg")
    assert out[:4] == b"%PDF"
