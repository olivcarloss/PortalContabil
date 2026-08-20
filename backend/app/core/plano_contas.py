import unicodedata
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook

_TIPOS_VALIDOS = {"ativo", "passivo", "receita", "custo", "despesa", "patrimonio_liquido"}
_TIPO_ALIASES = {
    "patrimonio liquido": "patrimonio_liquido",
    "patrimonio_liquido": "patrimonio_liquido",
    "pl": "patrimonio_liquido",
}

# Cada chave interna aceita varios nomes de coluna comuns em planilhas de
# plano de contas brasileiras — o layout real de cada escritorio varia.
_COLUNAS_ACEITAS: dict[str, set[str]] = {
    "codigo": {"codigo", "cod", "cod conta", "codigo conta", "codigo da conta", "conta codigo"},
    "descricao": {
        "descricao",
        "descricao da conta",
        "nome",
        "nome da conta",
        "conta",
        "conta contabil",
        "titulo",
    },
    "tipo": {
        "tipo",
        "tipo de conta",
        "tipo conta",
        "natureza",
        "classificacao",
        "classe",
        "grupo",
        "categoria",
    },
}


def _normalizar(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return " ".join(texto.strip().lower().split())


def _detectar_header(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    """Procura, nas primeiras linhas da planilha, aquela que contem pelo
    menos 2 das 3 colunas esperadas (codigo/descricao/tipo) — cobre o caso
    de uma linha de titulo acima do cabecalho real."""
    melhor: tuple[int, dict[str, int]] | None = None
    for linha_idx, row in enumerate(rows[:5]):
        header = [_normalizar(str(c)) if c is not None else "" for c in row]
        col_index: dict[str, int] = {}
        for chave, aliases in _COLUNAS_ACEITAS.items():
            for i, nome_coluna in enumerate(header):
                if nome_coluna in aliases:
                    col_index[chave] = i
                    break
        if len(col_index) >= 2 and (melhor is None or len(col_index) > len(melhor[1])):
            melhor = (linha_idx, col_index)
        if len(col_index) == 3:
            return melhor
    if melhor is not None:
        return melhor
    return 0, {}


def parse_plano_contas_xlsx(conteudo: bytes) -> list[dict]:
    """Le um XLS de plano de contas com colunas de codigo / descricao / tipo
    (nomes de coluna flexiveis — ver _COLUNAS_ACEITAS), procurando a linha
    de cabecalho entre as 5 primeiras linhas da planilha. Tipo aceita
    ativo/passivo/receita/despesa/patrimonio_liquido, ou sinonimos como
    'Patrimonio Liquido'/'PL'."""
    try:
        wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - arquivo invalido, converte pra 422
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Nao foi possivel ler o arquivo XLS: {exc}",
        ) from exc

    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="Planilha vazia")

    header_idx, col_index = _detectar_header(rows)
    faltando = set(_COLUNAS_ACEITAS) - set(col_index)
    if faltando:
        cabecalho_encontrado = [str(c) for c in rows[header_idx] if c is not None]
        raise HTTPException(
            status_code=422,
            detail=(
                "Colunas obrigatorias nao reconhecidas no XLS: "
                f"{', '.join(sorted(faltando))}. "
                f"Cabecalho lido na planilha: {cabecalho_encontrado or '(vazio)'}. "
                "Renomeie a coluna correspondente para Codigo, Descricao ou Tipo "
                "(ou um sinonimo comum, como Natureza/Classificacao para Tipo)."
            ),
        )

    contas: list[dict] = []
    for linha_num, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        codigo = row[col_index["codigo"]] if col_index["codigo"] < len(row) else None
        descricao = row[col_index["descricao"]] if col_index["descricao"] < len(row) else None
        tipo_raw = row[col_index["tipo"]] if col_index["tipo"] < len(row) else None
        if codigo is None and descricao is None and tipo_raw is None:
            continue
        if codigo is None or descricao is None or tipo_raw is None:
            raise HTTPException(
                status_code=422,
                detail=f"Linha {linha_num}: codigo, descricao e tipo sao obrigatorios",
            )
        tipo_norm = _normalizar(str(tipo_raw))
        tipo = _TIPO_ALIASES.get(tipo_norm, tipo_norm)
        if tipo not in _TIPOS_VALIDOS:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Linha {linha_num}: tipo \"{tipo_raw}\" invalido. "
                    "Use Ativo, Passivo, Receita, Custo, Despesa ou Patrimonio Liquido."
                ),
            )
        contas.append(
            {
                "codigo": str(codigo).strip(),
                "descricao": str(descricao).strip(),
                "tipo": tipo,
            }
        )

    if not contas:
        raise HTTPException(status_code=422, detail="Nenhuma conta encontrada no XLS")

    codigos_vistos = set()
    for conta in contas:
        if conta["codigo"] in codigos_vistos:
            raise HTTPException(
                status_code=422,
                detail=f"Codigo de conta duplicado no arquivo: {conta['codigo']}",
            )
        codigos_vistos.add(conta["codigo"])

    return contas
